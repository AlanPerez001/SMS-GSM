import concurrent.futures
import datetime
import glob
import logging
import sched
import sqlite3
import sys
import threading
import time
from datetime import datetime
import pkg_resources.py2_warn
import openpyxl
import serial
import serial.tools.list_ports
import xlsxwriter
from gsmmodem.modem import GsmModem
import PyQt5
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog, QMessageBox

conected = False
db_conn = ''
COM5 = 'Is_Desconnected'
COM6 = 'Is_Desconnected'
COM7 = 'Is_Desconnected'
COM8 = 'Is_Desconnected'
COM9 = 'Is_Desconnected'
COM10 = 'Is_Desconnected'
COM11 = 'Is_Desconnected'
COM12 = 'Is_Desconnected'

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setEnabled(True)
        MainWindow.resize(1048, 702)
        MainWindow.setMaximumSize(QtCore.QSize(1048, 702))
        MainWindow.setDocumentMode(False)
        MainWindow.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.treeWidget = QtWidgets.QTreeWidget(self.centralwidget)
        self.treeWidget.setGeometry(QtCore.QRect(0, 50, 221, 341))
        self.treeWidget.setObjectName("treeWidget")
        item_0 = QtWidgets.QTreeWidgetItem(self.treeWidget)
        item_0 = QtWidgets.QTreeWidgetItem(self.treeWidget)
        item_0 = QtWidgets.QTreeWidgetItem(self.treeWidget)
        item_0 = QtWidgets.QTreeWidgetItem(self.treeWidget)
        self.tableWidget = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(-10, 390, 231, 291))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Ignored, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tableWidget.sizePolicy().hasHeightForWidth())
        self.tableWidget.setSizePolicy(sizePolicy)
        self.tableWidget.setMidLineWidth(1)
        self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
        self.tableWidget.setAutoScrollMargin(10)
        self.tableWidget.setGridStyle(QtCore.Qt.NoPen)
        self.tableWidget.setRowCount(13)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(3)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(2, item)
        self.tableWidget.horizontalHeader().setVisible(True)
        self.tableWidget.horizontalHeader().setDefaultSectionSize(69)
        self.tableWidget.horizontalHeader().setMinimumSectionSize(30)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.verticalHeader().setDefaultSectionSize(25)
        self.tableWidget.verticalHeader().setMinimumSectionSize(23)
        self.stackedWidget = QtWidgets.QStackedWidget(self.centralwidget)
        self.stackedWidget.setEnabled(True)
        self.stackedWidget.setGeometry(QtCore.QRect(220, 50, 831, 631))
        self.stackedWidget.setAutoFillBackground(False)
        self.stackedWidget.setFrameShape(QtWidgets.QFrame.Panel)
        self.stackedWidget.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.stackedWidget.setObjectName("stackedWidget")
        self.page = QtWidgets.QWidget()
        self.page.setObjectName("page")
        self.pushButton = QtWidgets.QPushButton(self.page)
        self.pushButton.setGeometry(QtCore.QRect(370, 490, 111, 41))
        self.pushButton.setObjectName("pushButton")
        self.label = QtWidgets.QLabel(self.page)
        self.label.setGeometry(QtCore.QRect(210, 80,400, 41))
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiBold")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.Num_edit = QtWidgets.QPlainTextEdit(self.page)
        self.Num_edit.setGeometry(QtCore.QRect(300, 130, 241, 31))
        self.Num_edit.setObjectName("Num_edit")
        self.label_2 = QtWidgets.QLabel(self.page)
        self.label_2.setGeometry(QtCore.QRect(330, 210, 180, 31))
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiBold")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.SMSText_edit = QtWidgets.QPlainTextEdit(self.page)
        self.SMSText_edit.setGeometry(QtCore.QRect(240, 270, 331, 181))
        self.SMSText_edit.setObjectName("SMSText_edit")
        self.label.raise_()
        self.Num_edit.raise_()
        self.label_2.raise_()
        self.SMSText_edit.raise_()
        self.pushButton.raise_()
        self.stackedWidget.addWidget(self.page)
        self.page_2 = QtWidgets.QWidget()
        self.page_2.setObjectName("page_2")
        self.label_3 = QtWidgets.QLabel(self.page_2)
        self.label_3.setGeometry(QtCore.QRect(200, 40, 450, 19))
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiBold")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.pushButton_5 = QtWidgets.QPushButton(self.page_2)
        self.pushButton_5.setGeometry(QtCore.QRect(320, 400, 161, 51))
        self.pushButton_5.setObjectName("pushButton_5")
        self.listWidget = QtWidgets.QListWidget(self.page_2)
        self.listWidget.setGeometry(QtCore.QRect(260, 90, 256, 192))
        self.listWidget.setObjectName("listWidget")
        self.Actualizar = QtWidgets.QPushButton(self.page_2)
        self.Actualizar.setGeometry(QtCore.QRect(330, 550, 141, 41))
        self.Actualizar.setObjectName("Actualizar")
        self.progressBar = QtWidgets.QProgressBar(self.page_2)
        self.progressBar.setEnabled(True)
        self.progressBar.setGeometry(QtCore.QRect(300, 330, 201, 21))
        self.progressBar.setProperty("value", 0)
        self.progressBar.setAlignment(QtCore.Qt.AlignCenter)
        self.progressBar.setTextVisible(True)
        self.progressBar.setObjectName("progressBar")
        self.stackedWidget.addWidget(self.page_2)
        self.page_3 = QtWidgets.QWidget()
        self.page_3.setObjectName("page_3")
        self.dateEdit = QtWidgets.QDateEdit(self.page_3)
        self.dateEdit.setGeometry(QtCore.QRect(270, 170, 141, 31))
        self.dateEdit.setObjectName("dateEdit")
        self.label_4 = QtWidgets.QLabel(self.page_3)
        self.label_4.setGeometry(QtCore.QRect(200, 30, 550, 41))
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiBold")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.timeEdit = QtWidgets.QTimeEdit(self.page_3)
        self.timeEdit.setGeometry(QtCore.QRect(440, 170, 131, 31))
        self.timeEdit.setObjectName("timeEdit")
        self.label_5 = QtWidgets.QLabel(self.page_3)
        self.label_5.setGeometry(QtCore.QRect(270, 140, 51, 21))
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiBold")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.page_3)
        self.label_6.setGeometry(QtCore.QRect(440, 140, 51, 21))
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiBold")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.label_8 = QtWidgets.QLabel(self.page_3)
        self.label_8.setGeometry(QtCore.QRect(155, 280, 95, 21))
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiBold")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(self.page_3)
        self.label_9.setGeometry(QtCore.QRect(160, 360, 81, 21))
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiBold")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.num_edit_sched = QtWidgets.QPlainTextEdit(self.page_3)
        self.num_edit_sched.setGeometry(QtCore.QRect(270, 280, 151, 31))
        self.num_edit_sched.setObjectName("num_edit_sched")
        self.text_edit_sched = QtWidgets.QPlainTextEdit(self.page_3)
        self.text_edit_sched.setGeometry(QtCore.QRect(270, 360, 341, 161))
        self.text_edit_sched.setObjectName("text_edit_sched")
        self.Programa_btn = QtWidgets.QPushButton(self.page_3)
        self.Programa_btn.setGeometry(QtCore.QRect(390, 540, 111, 51))
        self.Programa_btn.setObjectName("Programa_btn")
        self.stackedWidget.addWidget(self.page_3)
        self.page_4 = QtWidgets.QWidget()
        self.page_4.setObjectName("page_4")
        self.tableWidget_3 = QtWidgets.QTableWidget(self.page_4)
        self.tableWidget_3.setGeometry(QtCore.QRect(0, 0, 831, 641))
        self.tableWidget_3.setFocusPolicy(QtCore.Qt.NoFocus)
        self.tableWidget_3.setAutoFillBackground(True)
        self.tableWidget_3.setAutoScroll(True)
        self.tableWidget_3.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.tableWidget_3.setShowGrid(True)
        self.tableWidget_3.setWordWrap(True)
        self.tableWidget_3.setCornerButtonEnabled(True)
        self.tableWidget_3.setRowCount(500)
        self.tableWidget_3.setObjectName("tableWidget_3")
        self.tableWidget_3.setColumnCount(4)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setItem(0, 0, item)
        self.tableWidget_3.horizontalHeader().setDefaultSectionSize(163)
        self.tableWidget_3.horizontalHeader().setMinimumSectionSize(40)
        self.tableWidget_3.verticalHeader().setVisible(False)
        self.stackedWidget.addWidget(self.page_4)
        self.page_5 = QtWidgets.QWidget()
        self.page_5.setObjectName("page_5")
        self.tableWidget_2 = QtWidgets.QTableWidget(self.page_5)
        self.tableWidget_2.setEnabled(True)
        self.tableWidget_2.setGeometry(QtCore.QRect(0, 0, 831, 631))
        self.tableWidget_2.setFocusPolicy(QtCore.Qt.NoFocus)
        self.tableWidget_2.setContextMenuPolicy(QtCore.Qt.DefaultContextMenu)
        self.tableWidget_2.setRowCount(500)
        self.tableWidget_2.setObjectName("tableWidget_2")
        self.tableWidget_2.setColumnCount(3)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(10, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(11, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(12, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(13, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(14, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(15, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(16, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(17, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(18, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(19, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(20, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(21, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(22, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setVerticalHeaderItem(23, item)
        item = QtWidgets.QTableWidgetItem()
        item.setText("Numero")
        self.tableWidget_2.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(2, item)
        self.tableWidget_2.horizontalHeader().setDefaultSectionSize(268)
        self.stackedWidget.addWidget(self.page_5)
        self.page_8 = QtWidgets.QWidget()
        self.page_8.setObjectName("page_8")
        self.tableWidget_4 = QtWidgets.QTableWidget(self.page_8)
        self.tableWidget_4.setGeometry(QtCore.QRect(0, 0, 831, 541))
        self.tableWidget_4.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.tableWidget_4.setRowCount(500)
        self.tableWidget_4.setObjectName("tableWidget_4")
        self.tableWidget_4.setColumnCount(3)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setVerticalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setItem(0, 0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setItem(1, 0, item)
        self.tableWidget_4.horizontalHeader().setDefaultSectionSize(276)
        self.tableWidget_4.verticalHeader().setVisible(False)
        self.Enviar = QtWidgets.QPushButton(self.page_8)
        self.Enviar.setGeometry(QtCore.QRect(350, 560, 131, 41))
        self.Enviar.setObjectName("Enviar")
        self.Actualiza_db = QtWidgets.QPushButton(self.page_8)
        self.Actualiza_db.setGeometry(QtCore.QRect(20, 570, 81, 31))
        self.Actualiza_db.setObjectName("Actualiza_db")
        self.stackedWidget.addWidget(self.page_8)
        self.page_6 = QtWidgets.QWidget()
        self.page_6.setObjectName("page_6")
        self.Create_DB = QtWidgets.QPushButton(self.page_6)
        self.Create_DB.setGeometry(QtCore.QRect(560, 220, 81, 31))
        self.Create_DB.setObjectName("Create_DB")
        self.Name_create_db = QtWidgets.QPlainTextEdit(self.page_6)
        self.Name_create_db.setGeometry(QtCore.QRect(320, 220, 211, 31))
        self.Name_create_db.setObjectName("Name_create_db")
        self.label_10 = QtWidgets.QLabel(self.page_6)
        self.label_10.setGeometry(QtCore.QRect(85, 220, 250, 16))
        self.label_10.setObjectName("label_10")
        self.Importar_db = QtWidgets.QPushButton(self.page_6)
        self.Importar_db.setGeometry(QtCore.QRect(370, 380, 111, 41))
        self.Importar_db.setObjectName("Importar_db")
        self.stackedWidget.addWidget(self.page_6)
        self.page_7 = QtWidgets.QWidget()
        self.page_7.setObjectName("page_7")
        self.rep_enviados_checkbx = QtWidgets.QCheckBox(self.page_7)
        self.rep_enviados_checkbx.setGeometry(QtCore.QRect(150, 190, 80, 16))
        self.rep_enviados_checkbx.setObjectName("rep_enviados_checkbx")
        self.rep_recib_checkbx = QtWidgets.QCheckBox(self.page_7)
        self.rep_recib_checkbx.setGeometry(QtCore.QRect(150, 280, 85, 16))
        self.rep_recib_checkbx.setObjectName("rep_recib_checkbx")
        self.rep_todo_checkbx = QtWidgets.QCheckBox(self.page_7)
        self.rep_todo_checkbx.setGeometry(QtCore.QRect(150, 110, 75, 16))
        self.rep_todo_checkbx.setChecked(True)
        self.rep_todo_checkbx.setObjectName("rep_todo_checkbx")
        self.genera_rep_btn = QtWidgets.QPushButton(self.page_7)
        self.genera_rep_btn.setGeometry(QtCore.QRect(290, 420, 141, 41))
        self.genera_rep_btn.setObjectName("genera_rep_btn")
        self.stackedWidget.addWidget(self.page_7)
        self.IniciarBTN = QtWidgets.QPushButton(self.centralwidget)
        self.IniciarBTN.setGeometry(QtCore.QRect(10, 10, 121, 31))
        self.IniciarBTN.setObjectName("IniciarBTN")
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(160, 10, 121, 31))
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.PuertosBTN = QtWidgets.QPushButton(self.centralwidget)
        self.PuertosBTN.setGeometry(QtCore.QRect(310, 10, 121, 31))
        self.PuertosBTN.setObjectName("PuertosBTN")
        self.ExitBTN = QtWidgets.QPushButton(self.centralwidget)
        self.ExitBTN.setGeometry(QtCore.QRect(930, 10, 101, 31))
        self.ExitBTN.setObjectName("ExitBTN")
        self.Connect_DB = QtWidgets.QPushButton(self.centralwidget)
        self.Connect_DB.setGeometry(QtCore.QRect(460, 10, 141, 31))
        self.Connect_DB.setObjectName("Connect_DB")
        self.crear_campaa = QtWidgets.QPushButton(self.centralwidget)
        self.crear_campaa.setGeometry(QtCore.QRect(630, 10, 121, 31))
        self.crear_campaa.setObjectName("crear_campaa")
        self.Reporte_camp = QtWidgets.QPushButton(self.centralwidget)
        self.Reporte_camp.setGeometry(QtCore.QRect(765, 10, 150, 31))
        self.Reporte_camp.setObjectName("Reporte_camp")
        self.Stop_btn = QtWidgets.QPushButton(self.centralwidget)
        self.Stop_btn.setGeometry(QtCore.QRect(10, 10, 121, 31))
        self.Stop_btn.setObjectName("Stop_btn")
        self.line = QtWidgets.QFrame(self.centralwidget)
        self.line.setGeometry(QtCore.QRect(910, 0, 20, 51))
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.line_2 = QtWidgets.QFrame(self.centralwidget)
        self.line_2.setGeometry(QtCore.QRect(130, 0, 20, 51))
        self.line_2.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.comboBox.raise_()
        self.treeWidget.raise_()
        self.tableWidget.raise_()
        self.stackedWidget.raise_()
        self.IniciarBTN.raise_()
        self.PuertosBTN.raise_()
        self.ExitBTN.raise_()
        self.Connect_DB.raise_()
        self.crear_campaa.raise_()
        self.Reporte_camp.raise_()
        self.Stop_btn.raise_()
        self.line.raise_()
        self.line_2.raise_()
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1048, 21))
        self.menubar.setObjectName("menubar")
        self.menuArchivo = QtWidgets.QMenu(self.menubar)
        self.menuArchivo.setObjectName("menuArchivo")
        self.menuHerramientas = QtWidgets.QMenu(self.menubar)
        self.menuHerramientas.setObjectName("menuHerramientas")
        self.menuAyuda = QtWidgets.QMenu(self.menubar)
        self.menuAyuda.setObjectName("menuAyuda")
        MainWindow.setMenuBar(self.menubar)
        self.actionConectarModem = QtWidgets.QAction(MainWindow)
        self.actionConectarModem.setObjectName("actionConectarModem")
        self.actionEnviar_mensaje = QtWidgets.QAction(MainWindow)
        self.actionEnviar_mensaje.setObjectName("actionEnviar_mensaje")
        self.actionProgramar_mensaje = QtWidgets.QAction(MainWindow)
        self.actionProgramar_mensaje.setObjectName("actionProgramar_mensaje")
        self.actionExit = QtWidgets.QAction(MainWindow)
        self.actionExit.setObjectName("actionExit")
        self.actionIniciar = QtWidgets.QAction(MainWindow)
        self.actionIniciar.setObjectName("actionIniciar")
        self.actionPuertos = QtWidgets.QAction(MainWindow)
        self.actionPuertos.setObjectName("actionPuertos")
        self.menuArchivo.addAction(self.actionConectarModem)
        self.menuArchivo.addAction(self.actionEnviar_mensaje)
        self.menuArchivo.addAction(self.actionProgramar_mensaje)
        self.menuArchivo.addSeparator()
        self.menuArchivo.addAction(self.actionExit)
        self.menuHerramientas.addAction(self.actionIniciar)
        self.menuHerramientas.addAction(self.actionPuertos)
        self.menubar.addAction(self.menuArchivo.menuAction())
        self.menubar.addAction(self.menuHerramientas.menuAction())
        self.menubar.addAction(self.menuAyuda.menuAction())

        self.retranslateUi(MainWindow)
        self.stackedWidget.setCurrentIndex(6)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Sms Sender"))
        self.treeWidget.headerItem().setText(0, _translate("MainWindow", "Menu"))
        __sortingEnabled = self.treeWidget.isSortingEnabled()
        self.treeWidget.setSortingEnabled(False)
        self.treeWidget.topLevelItem(0).setText(0, _translate("MainWindow", "Enviados"))
        self.treeWidget.topLevelItem(1).setText(0, _translate("MainWindow", "Recibidos"))
        self.treeWidget.topLevelItem(2).setText(0, _translate("MainWindow", "Programados"))
        self.treeWidget.topLevelItem(3).setText(0, _translate("MainWindow", "Base de datos"))
        self.treeWidget.setSortingEnabled(__sortingEnabled)
        self.tableWidget.setSortingEnabled(False)
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "No."))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Puerto"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Estado"))
        self.pushButton.setText(_translate("MainWindow", "Enviar"))
        self.label.setText(_translate("MainWindow", "Ingrese el Numero destinatario:"))
        self.label_2.setText(_translate("MainWindow", "Escriba el mensaje:"))
        self.label_3.setText(_translate("MainWindow", "Selecciona uno de los puertos disponibles:"))
        self.pushButton_5.setText(_translate("MainWindow", "Conectar"))
        self.Actualizar.setText(_translate("MainWindow", "Actualizar Puertos"))
        self.label_4.setText(_translate("MainWindow", "Selecciona el día y la hora para el envio del mensaje:"))
        self.label_5.setText(_translate("MainWindow", "Fecha:"))
        self.label_6.setText(_translate("MainWindow", "Hora:"))
        self.label_8.setText(_translate("MainWindow", "Destinatario:"))
        self.label_9.setText(_translate("MainWindow", "Mensaje:"))
        self.Programa_btn.setText(_translate("MainWindow", "Prgramar"))
        self.tableWidget_3.setSortingEnabled(False)
        item = self.tableWidget_3.verticalHeaderItem(0)
        item.setText(_translate("MainWindow", "1"))
        item = self.tableWidget_3.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Numero"))
        item = self.tableWidget_3.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Mensaje"))
        item = self.tableWidget_3.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Puerto"))
        item = self.tableWidget_3.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "Hora de envio"))
        __sortingEnabled = self.tableWidget_3.isSortingEnabled()
        self.tableWidget_3.setSortingEnabled(False)
        self.tableWidget_3.setSortingEnabled(__sortingEnabled)
        item = self.tableWidget_2.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Mensaje"))
        item = self.tableWidget_2.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Hore de recibido"))
        item = self.tableWidget_4.verticalHeaderItem(0)
        item.setText(_translate("MainWindow", "New Row"))
        item = self.tableWidget_4.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Numero"))
        item = self.tableWidget_4.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Mensaje"))
        item = self.tableWidget_4.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Enviado"))
        __sortingEnabled = self.tableWidget_4.isSortingEnabled()
        self.tableWidget_4.setSortingEnabled(False)
        self.tableWidget_4.setSortingEnabled(__sortingEnabled)
        self.Enviar.setText(_translate("MainWindow", "Enviar"))
        self.Actualiza_db.setText(_translate("MainWindow", "Actualiza"))
        self.Create_DB.setText(_translate("MainWindow", "Crear"))
        self.label_10.setText(_translate("MainWindow", "Introduce el nombre de la capaña:"))
        self.Importar_db.setText(_translate("MainWindow", "Importar"))
        self.rep_enviados_checkbx.setText(_translate("MainWindow", "Enviados"))
        self.rep_recib_checkbx.setText(_translate("MainWindow", "Recibidos"))
        self.rep_todo_checkbx.setText(_translate("MainWindow", "Todo"))
        self.genera_rep_btn.setText(_translate("MainWindow", "Generar reporte"))
        self.IniciarBTN.setText(_translate("MainWindow", "Iniciar"))
        self.comboBox.setItemText(0, _translate("MainWindow", "Enviar sms"))
        self.comboBox.setItemText(1, _translate("MainWindow", "Componer"))
        self.comboBox.setItemText(2, _translate("MainWindow", "Programado"))
        self.PuertosBTN.setText(_translate("MainWindow", "Puertos"))
        self.ExitBTN.setText(_translate("MainWindow", "Salir"))
        self.Connect_DB.setText(_translate("MainWindow", "Conectar campaña"))
        self.crear_campaa.setText(_translate("MainWindow", "Crear campaña"))
        self.Reporte_camp.setText(_translate("MainWindow", "Reporte de campaña"))
        self.Stop_btn.setText(_translate("MainWindow", "Detener"))
        self.menuArchivo.setTitle(_translate("MainWindow", "Archivo"))
        self.menuHerramientas.setTitle(_translate("MainWindow", "Herramientas"))
        self.menuAyuda.setTitle(_translate("MainWindow", "Ayuda"))
        self.actionConectarModem.setText(_translate("MainWindow", "Conectar"))
        self.actionEnviar_mensaje.setText(_translate("MainWindow", "Enviar mensaje"))
        self.actionProgramar_mensaje.setText(_translate("MainWindow", "Programar mensaje"))
        self.actionExit.setText(_translate("MainWindow", "Exit"))
        self.actionIniciar.setText(_translate("MainWindow", "Iniciar"))
        self.actionPuertos.setText(_translate("MainWindow", "Puertos"))

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):

    # Inicia aplicacion y actionbuttons que activan las funciones
    def __init__(self, *args, **kwargs):
        global modem5
        global modem6
        global modem7
        global modem8
        global modem9
        global modem10
        global modem11
        global modem12
        modem5 = GsmModem(
            'COM5', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem6 = GsmModem(
            'COM6', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem7 = GsmModem(
            'COM7', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem8 = GsmModem(
            'COM8', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem9 = GsmModem(
            'COM9', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem10 = GsmModem(
            'COM10', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem11 = GsmModem(
            'COM11', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem12 = GsmModem(
            'COM12', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        QtWidgets.QMainWindow.__init__(self, *args, **kwargs)
        self.setupUi(self)
        self.dateEdit.setDateTime(QtCore.QDateTime.currentDateTime())
        self.dateEdit.setCalendarPopup(True)
        self.stackedWidget.setCurrentIndex(1)
        self.progressBar.hide()
        self.Stop_btn.hide()
        self.treeWidget.itemClicked.connect(self.MenuSend)
        self.actionConectarModem.triggered.connect(self.actionConnect)
        self.actionPuertos.triggered.connect(self.actionConnect)
        self.actionIniciar.triggered.connect(self.startcon)
        self.actionEnviar_mensaje.triggered.connect(self.actionSend)
        self.actionProgramar_mensaje.triggered.connect(self.actionSchedul)
        self.actionExit.triggered.connect(self.actionExitapp)
        self.pushButton_5.clicked.connect(self.ConectModem)
        self.Actualizar.clicked.connect(self.ActualizarPort)
        self.ExitBTN.clicked.connect(self.actionExitapp)
        self.IniciarBTN.clicked.connect(self.startcon)
        self.Stop_btn.clicked.connect(self.stopcon)
        self.Enviar.clicked.connect(self.start_send)
        self.comboBox.currentIndexChanged.connect(self.combo)
        self.PuertosBTN.clicked.connect(self.actionConnect)
        self.Connect_DB.clicked.connect(self.getfile)
        self.Actualiza_db.clicked.connect(self.updateDB)
        self.crear_campaa.clicked.connect(self.create_db)
        self.Reporte_camp.clicked.connect(self.actionReporte)
        self.genera_rep_btn.clicked.connect(self.generar_reporte)
        self.Programa_btn.clicked.connect(self.schedule_sms)
        
   # Inicia la coneccion con todos los puertos disponibles
    def startcon(self):
        global conected
        hilos_conn = threading.Thread(
            target=self.hilos, args=('Connect_device',))
        hilos_conn.start()
        hilos_conn.join()
        conected = True
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setInformativeText('Conneccion establecida con los puertos :D')
        msg.setWindowTitle("info")
        msg.exec_()
        self.Stop_btn.show()

    #Detiene la coneccion con los puertos
    def stopcon(self):
        self.hilos('Stop_device')
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setInformativeText('Puertos desconectados')
        msg.setWindowTitle("info")
        msg.exec_()
        self.Stop_btn.hide()
    
    # Muestra el index 6
    def create_show(self):
        self.stackedWidget.setCurrentIndex(6)
        self.Importar_db.clicked.connect(self.import_db)

    # Lee los datos en el excel del usuario para insertarlos a la base de datos
    def import_db(self):
        excel_file, _filter = QFileDialog.getOpenFileName(self, 'Selecciona tu archivo XLSX a importar',
                                                          'C:\\Users\\Developer\\Desktop\\', "Excel File (*.xlsx)")
        if excel_file != '':
            wb_obj = openpyxl.load_workbook(excel_file)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            con = sqlite3.connect(db_conn)
            cursor = con.cursor()
            for i in range(2, m_row + 1):
                numero = sheet_obj.cell(row=i, column=1)
                mensaje = sheet_obj.cell(row=i, column=2)
                cursor.execute('INSERT INTO SMS (Numero,Mensaje,Enviado,Intentos) VALUES("' +
                            str(numero.value)+'","'+str(mensaje.value)+'","F",0)')
                con.commit()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setInformativeText('Registros importados a la base de datos')
            msg.setWindowTitle("info")
            msg.exec_()
            self.updateDB()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setInformativeText(
                'No se selecciono ningun archivo a importar')
            msg.setWindowTitle("info")
            msg.exec_()

    # crea una base de datos nueva con el nombre escrito por el usuario en la ubicacion seleccionada
    def create_db(self):
        global db_conn
        filename = QFileDialog.getSaveFileName(self, 'Crear campaña', '', "Data Base file (*.db);;All Files (*)")
        database= str(filename[0])
        sms = """ CREATE TABLE "SMS" (
                    "ID"	INTEGER UNIQUE,
                    "Numero"	INTEGER,
                    "Mensaje"	TEXT,
                    "Enviado"	TEXT,
                    "Intentos"	INTEGER,
                    PRIMARY KEY("ID" AUTOINCREMENT)
                );"""
        enviados = """CREATE TABLE "Enviados" (
                            "ID"	INTEGER,
                            "Numero"	INTEGER,
                            "Mensaje"	TEXT,
                            "Puerto"	TEXT,
                            "Hora_envio"	TEXT,
                            "Hora_ingreso_db"	TEXT,
                            PRIMARY KEY("ID" AUTOINCREMENT)
                        );"""
        recibidos = """CREATE TABLE "Recibidos" (
                            "ID"	INTEGER UNIQUE,
                            "Numero"	INTEGER,
                            "Mensaje"	TEXT,
                            "Hora_recibido"	TEXT,
                            PRIMARY KEY("ID" AUTOINCREMENT)
                        );"""
        individuales = """CREATE TABLE "Individuales" (
                            "ID"	INTEGER,
                            "Numero"	INTEGER,
                            "Mensaje"	TEXT,
                            PRIMARY KEY("ID" AUTOINCREMENT)
                        );"""
        conn = sqlite3.connect(database)
        try:
            c = conn.cursor()
            c.execute(sms)
        except Exception as e:
            print(e)
        try:
            c = conn.cursor()
            c.execute(enviados)
        except Exception as e:
            print(e)
        try:
            c = conn.cursor()
            c.execute(individuales)
        except Exception as e:
            print(e)
        try:
            c = conn.cursor()
            c.execute(recibidos)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setInformativeText('Base de datos creada con exito :D')
            msg.setWindowTitle("info")
            msg.exec_()
            self.Importar_db.show()
            db_conn = database
            self.import_db()
        except Exception as e:
            print(e)

    # Obtiene la ruta de la base de datos seleccionada
    def getfile(self):
        global db_conn
        db_conn, _filter = QFileDialog.getOpenFileName(self, 'Selecciona tu base de datos',
                                                       'C:\\Users\\Developer\\Desktop\\', "data base file (*);;All Files (*)")
        self.updateDB()

    #Muestra index 7
    def actionReporte(self):
        self.stackedWidget.setCurrentIndex(7)
        checked = self.rep_recib_checkbx.isChecked()
    
    #Generar reporte de la campaña
    def generar_reporte(self):
        if db_conn != '':
            filename = QFileDialog.getSaveFileName(self, 'Guarda tu archivo', '', "Excel Files (*.xlsx);;All Files (*)")
            if filename[0] != '':
                workbook = xlsxwriter.Workbook(filename[0]) 
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                recib_check = self.rep_recib_checkbx.isChecked()
                env_check = self.rep_enviados_checkbx.isChecked()
                todo_check = self.rep_todo_checkbx.isChecked()
                if todo_check == True:
                    recib_check = True
                    env_check = True
                if recib_check == True:
                    worksheet = workbook.add_worksheet('Recibidos') 
                    cursor.execute('SELECT * FROM Recibidos')
                    query = cursor.fetchall()
                    worksheet.write(0,0,'ID')
                    worksheet.write(0,1,'Numero')
                    worksheet.write(0,2,'Mensaje')
                    worksheet.write(0,3,'Hora de recibido')
                    worksheet.write(0,4,'Hora de insert a la base')
                    row = 1
                    col = 0
                    for pk,num,mnsj,rec in query:
                        worksheet.write(row, col, pk) 
                        worksheet.write(row, col+1, num)
                        worksheet.write(row, col+2, mnsj)
                        worksheet.write(row, col+3, rec)
                        row +=1
                if env_check == True:
                    worksheet = workbook.add_worksheet('Enviados') 
                    cursor.execute('SELECT * FROM Enviados')
                    query = cursor.fetchall()
                    worksheet.write(0,0,'ID')
                    worksheet.write(0,1,'Numero')
                    worksheet.write(0,2,'Mensaje')
                    worksheet.write(0,3,'Puerto')
                    worksheet.write(0,4,'Hora de Enviado')
                    row = 1
                    col = 0
                    for pk,num,mnsj,puerto,hora,hora_insert in query:
                        worksheet.write(row, col, pk) 
                        worksheet.write(row, col+1, num)
                        worksheet.write(row, col+2, mnsj)
                        worksheet.write(row, col+3, puerto)
                        worksheet.write(row, col+4, hora)
                        worksheet.write(row, col+5, hora_insert)
                        row +=1
                workbook.close()
            else:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setInformativeText(
                    'No se selecciono ninguna ruta :c')
                msg.setWindowTitle("info")
                msg.exec_()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setInformativeText(
                'No esta conectado a ninguna campaña')
            msg.setWindowTitle("info")
            msg.exec_()
    # muestra el index 2 mensajes programados (Coming Soon)
    def actionSchedul(self):
        self.stackedWidget.setCurrentIndex(2)

    # Cierra conecciones con los puertos abiertos y cierra la app
    def actionExitapp(self):

        try:
            self.stopcon()
        except Exception as e:
            print(e)
        finally:
            app.exit()

    # Muestra index 1
    def actionConnect(self):
        self.stackedWidget.setCurrentIndex(1)

    # Muestra index 0
    def actionSend(self):
        self.stackedWidget.setCurrentIndex(0)
        self.pushButton.clicked.connect(self.indivudual_sms)

    # Devuelve y activa el index seleccionado en el Combobox
    def combo(self):
        seleccion = self.comboBox.currentText()
        if str(seleccion) == 'Componer':
            self.stackedWidget.setCurrentIndex(0)
            self.pushButton.clicked.connect(self.indivudual_sms)
        elif str(seleccion) == 'Programado':
            self.stackedWidget.setCurrentIndex(2)

    # Establece una coneccion individual del puerto seleccionado
    def ConectModem(self):
        try:
            portSelected = self.listWidget.currentItem().text()
            try:
                self.modem = GsmModem(
                    str(portSelected), 115200, smsReceivedCallbackFunc=MainWindow.ReceivedSms)
                self.modem.smsTextMode = False
                self.modem.connect()
            except Exception as F:
                print('El puerto '+str(portSelected) +
                      ' No esta disponible: ' + str(F))
        except Exception as e:
            print(e)

    # obtiene una lista de los puertos existentes
    def ActualizarPort(self):
        try:
            ports = serial.tools.list_ports.comports()
            advance = 0
            porcantaje = float(100/len(ports))
            self.listWidget.clear()
            num_hilo = 1
            for p in sorted(ports, key=None, reverse=False):
                advance += float(porcantaje)
                port = p.device
                try:
                    self.progressBar.show()
                    self.progressBar.setGeometry(QtCore.QRect(300, 330, 201, 21))
                    self.progressBar.setProperty("value", advance)
                    self.progressBar.setObjectName("progressBar")
                    # Abre hilo llamando a la funcion Connectport para crear coneccion momentanea y corroborar el estado del puerto
                    hilo = threading.Thread(name='hilo %s' % str(
                        num_hilo), target=self.connectport, args=(port,))
                    hilo.start()
                    num_hilo += 1
                finally:
                    self.progressBar.hide()
        except Exception as e:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setInformativeText(
                    'No hay puertos disponibles')
                msg.setWindowTitle("Error")
                msg.exec_()


    # Crea una coneccion momentanea para corroborar el estado de los puertos y  escribe los puertos en la lista desplegable
    def connectport(self, port):
        try:
            phone = serial.Serial(str(port),  115200, timeout=0.2)
            phone.write(b'ATZ\r')
            phone.write(b'AT+CMGF=1\r')
            if str(phone.read()) != str(b''):
                self.listWidget.addItem(port)
        except Exception as E:
            if str(E) == 'None':
                print('No disponibe :c')
        finally:
            try:
                phone.close()
            except Exception as f:
                print(f)

    # Envia mensaje de texto al numero escrito por el usuario por el primer puerto disponible
    def indivudual_sms(self):
        num = self.Num_edit.toPlainText()
        text = self.SMSText_edit.toPlainText()
        send = False
        con = sqlite3.connect(str(db_conn))
        cursor = con.cursor()
        if conected == True:
            if num != "" and text != "" and len(num) >= 10:
                while send == False:
                    if send == False:
                        try:
                            if COM5 == 'Is_Connected':
                                date = []
                                modem5.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                date.append(str(time.asctime(time.localtime(time.time()))))
                                send = True
                                cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(num)+'","'+str(text)+'","COM5","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                                con.commit()
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                                self.updateDB()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM6 == 'Is_Connected':
                                modem6.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM6","'+str(time.asctime(time.localtime(time.time())))+'")')
                                con.commit()
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                                self.updateDB()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM7 == 'Is_Connected':
                                modem7.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM7","'+str(time.asctime(time.localtime(time.time())))+'")')
                                con.commit()
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                                self.updateDB()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM8 == 'Is_Connected':
                                modem8.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM8","'+str(time.asctime(time.localtime(time.time())))+'")')
                                con.commit()
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                                self.updateDB()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM9 == 'Is_Connected':
                                modem9.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM9","'+str(time.asctime(time.localtime(time.time())))+'")')
                                con.commit()
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                                self.updateDB()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM10 == 'Is_Connecte d':
                                modem10.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM10","'+str(time.asctime(time.localtime(time.time())))+'")')
                                con.commit()
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                                self.updateDB()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM11 == 'Is_Connected':
                                modem11.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM11","'+str(time.asctime(time.localtime(time.time())))+'")')
                                con.commit()
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                                self.updateDB()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM12 == 'Is_Connected':
                                modem12.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM12","'+str(time.asctime(time.localtime(time.time())))+'")')
                                con.commit()
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                                self.updateDB()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
            else:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setInformativeText(
                    'Por favor ingresa un numero y/o mesnsaje')
                msg.setWindowTitle("info")
                msg.exec_()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setInformativeText(
                'Por favor inicia los puertos para poder enviar mensajes')
            msg.setWindowTitle("Error")
            msg.exec_()


    def schedule_sms(self):
        date_form = self.dateEdit.date()
        hour_form = self.timeEdit.time()
        date=date_form.toPyDate()
        hour = hour_form.toPyTime()
        num = self.num_edit_sched.toPlainText()
        text = self.text_edit_sched.toPlainText()
        schedule = str(date)+'T'+str(hour)
        p = '%Y-%m-%dT%H:%M:%S'
        mytime = str(schedule)
        epoch = datetime(1970, 1, 1)
        time_sched = (datetime.strptime(mytime, p) - epoch).total_seconds()
        thread_sched = threading.Thread(target= self.thread_scheduled, args=(time_sched,num,text,))
        thread_sched.start()
        self.num_edit_sched.clear()
        self.text_edit_sched.clear()
    

    def thread_scheduled(self, tim,num,text):
        scheduler_e = sched.scheduler(time.time, time.sleep)
        print(tim+18000)
        scheduler_e.enterabs(int(tim+18000), 1, self.send_sched, (num,text))
        scheduler_e.run()
        

    def send_sched(self,num,text): 
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("Info")
        msg.setInformativeText('El mensaje programado se enviara ahora :D')
        msg.setWindowTitle("info")
        msg.exec_()
        con = sqlite3.connect(str(db_conn))
        cursor = con.cursor()
        send = False
        while send == False:
            if send == False:
                try:
                    if COM5 == 'Is_Connected':
                        date = []
                        modem5.sendSms(str(num), str(
                            text), waitForDeliveryReport=False, deliveryTimeout=1)
                        date.append(str(time.asctime(time.localtime(time.time()))))
                        send = True
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(num)+'","'+str(text)+'","COM5","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()

                        self.updateDB()
                except Exception as F:
                    send = False
                    if str(F) == 'CMS 500':
                        print('El puerto seleccionado devolvio ' +
                                str(F)+', [Error Unknown]')
            if send == False:
                try:
                    if COM6 == 'Is_Connected':
                        modem6.sendSms(str(num), str(
                            text), waitForDeliveryReport=False, deliveryTimeout=1)
                        send = True
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM6","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                        self.updateDB()
                except Exception as F:
                    send = False
                    if str(F) == 'CMS 500':
                        print('El puerto seleccionado devolvio ' +
                                str(F)+', [Error Unknown]')
            if send == False:
                try:
                    if COM7 == 'Is_Connected':
                        modem7.sendSms(str(num), str(
                            text), waitForDeliveryReport=False, deliveryTimeout=1)
                        send = True
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM7","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                        self.updateDB()
                except Exception as F:
                    send = False
                    if str(F) == 'CMS 500':
                        print('El puerto seleccionado devolvio ' +
                                str(F)+', [Error Unknown]')
            if send == False:
                try:
                    if COM8 == 'Is_Connected':
                        modem8.sendSms(str(num), str(
                            text), waitForDeliveryReport=False, deliveryTimeout=1)
                        send = True
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM8","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                        self.updateDB()
                except Exception as F:
                    send = False
                    if str(F) == 'CMS 500':
                        print('El puerto seleccionado devolvio ' +
                                str(F)+', [Error Unknown]')
            if send == False:
                try:
                    if COM9 == 'Is_Connected':
                        modem9.sendSms(str(num), str(
                            text), waitForDeliveryReport=False, deliveryTimeout=1)
                        send = True
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM9","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                        self.updateDB()
                except Exception as F:
                    send = False
                    if str(F) == 'CMS 500':
                        print('El puerto seleccionado devolvio ' +
                                str(F)+', [Error Unknown]')
            if send == False:
                try:
                    if COM10 == 'Is_Connecte d':
                        modem10.sendSms(str(num), str(
                            text), waitForDeliveryReport=False, deliveryTimeout=1)
                        send = True
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM10","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                        self.updateDB()
                except Exception as F:
                    send = False
                    if str(F) == 'CMS 500':
                        print('El puerto seleccionado devolvio ' +
                                str(F)+', [Error Unknown]')
            if send == False:
                try:
                    if COM11 == 'Is_Connected':
                        modem11.sendSms(str(num), str(
                            text), waitForDeliveryReport=False, deliveryTimeout=1)
                        send = True
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM11","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                        self.updateDB()
                except Exception as F:
                    send = False
                    if str(F) == 'CMS 500':
                        print('El puerto seleccionado devolvio ' +
                                str(F)+', [Error Unknown]')
            if send == False:
                try:
                    if COM12 == 'Is_Connected':
                        modem12.sendSms(str(num), str(
                            text), waitForDeliveryReport=False, deliveryTimeout=1)
                        send = True
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(num)+'","'+str(text)+'","COM12","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                        self.updateDB()
                except Exception as F:
                    send = False
                    if str(F) == 'CMS 500':
                        print('El puerto seleccionado devolvio ' +
                                str(F)+', [Error Unknown]')

    # Cuando un mensaje se recibe redirige a la funcion para obtener el mensaje
    def ReceivedSms(self, sms):
        con = sqlite3.connect(str(db_conn))
        cursor = con.cursor()
        cursor.execute('INSERT INTO Recibidos (Numero,Mensaje,Hora_recibido) VALUES("' +
                       str(sms.number)+'","'+str(sms.text)+'","'+str(sms.time)+'")')
        con.commit()
        self.updateDB()

    # Actualiza la base de datos Enviados y Mensajes
    def updateDB(self):
        if db_conn != '':
            con = sqlite3.connect(str(db_conn))
            cursor = con.cursor()
            cursor.execute('SELECT* FROM SMS')
            rows = cursor.fetchall()
            count = 0
            for i in rows:
                numero = QtWidgets.QTableWidgetItem()
                self.tableWidget_4.setItem(count, 0, numero)
                numero = self.tableWidget_4.item(count, 0)
                numero.setText(str(i[1]))
                mensaje = QtWidgets.QTableWidgetItem()
                self.tableWidget_4.setItem(count, 1, mensaje)
                mensaje = self.tableWidget_4.item(count, 1)
                mensaje.setText(str(i[2]))
                enviado = QtWidgets.QTableWidgetItem()
                self.tableWidget_4.setItem(count, 2, enviado)
                enviado = self.tableWidget_4.item(count, 2)
                enviado.setText(str(i[3]))
                count += 1

            cursor.execute('SELECT* FROM Enviados')
            rows = cursor.fetchall()
            count2 = 0
            for i in rows:
                numero = QtWidgets.QTableWidgetItem()
                self.tableWidget_3.setItem(count2, 0, numero)
                numero = self.tableWidget_3.item(count2, 0)
                numero.setText(str(i[1]))
                mensaje = QtWidgets.QTableWidgetItem()
                self.tableWidget_3.setItem(count2, 1, mensaje)
                mensaje = self.tableWidget_3.item(count2, 1)
                mensaje.setText(str(i[2]))
                com = QtWidgets.QTableWidgetItem()
                self.tableWidget_3.setItem(count2, 2, com)
                com = self.tableWidget_3.item(count2, 2)
                com.setText(str(i[3]))
                date = QtWidgets.QTableWidgetItem()
                self.tableWidget_3.setItem(count2, 3, date)
                date = self.tableWidget_3.item(count2, 3)
                date.setText(str(i[4]))
                count2 += 1

            cursor.execute('SELECT* FROM Recibidos')
            rows = cursor.fetchall()
            count3 = 0
            for i in rows:
                numero = QtWidgets.QTableWidgetItem()
                self.tableWidget_2.setItem(count3, 0, numero)
                numero = self.tableWidget_2.item(count3, 0)
                numero.setText(str(i[1]))
                mensaje = QtWidgets.QTableWidgetItem()
                self.tableWidget_2.setItem(count3, 1, mensaje)
                mensaje = self.tableWidget_2.item(count3, 1)
                mensaje.setText(str(i[2]))
                date = QtWidgets.QTableWidgetItem()
                self.tableWidget_2.setItem(count3, 2, date)
                date = self.tableWidget_2.item(count3, 2)
                date.setText(str(i[3]))
                count3 += 1
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Error")
            msg.setInformativeText(
                'Base de datos de campaña no seleccionados.')
            msg.setWindowTitle("Error")
            msg.exec_()

    # Puerto 5
    def Com5(self, action):
        global modem5
        global COM5
        port = 'COM5'
        if action == 'Connect_device':
            try:
                modem5.smsTextMode = False
                modem5.connect()
                COM5 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':
            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if parametro[4] < 2:
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        date=[]
                        modem5.sendSms(str(parametro[1]), str(parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        date.append(str(time.asctime(time.localtime(time.time()))))
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':

                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F",Intentos = '+str(int(parametro[4]+1))+' WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM5 
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
                else:   
                    cursor.execute('UPDATE SMS SET Enviado ="Error" WHERE ID ='+str(parametro[0]))
                    con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)
        elif action == 'Stop_device':
            try:
                modem5.close()
            except Exception as e:
                print(e)

    # Puerto 6
    def Com6(self, action):
        global modem6
        global COM6
        port = 'COM6'
        if action == 'Connect_device':
            try:
                modem6.smsTextMode = False
                modem6.connect()
                COM6 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':
            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if parametro[4] < 2:
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        date=[]
                        modem6.sendSms(str(parametro[1]), str(parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        date.append(str(time.asctime(time.localtime(time.time()))))
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':

                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F",Intentos = '+str(int(parametro[4]+1))+' WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM6 
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)
        elif action == 'Stop_device':
            try:
                modem6.close()
            except Exception as e:
                print(e)

    # Puerto 7
    def Com7(self, action):
        global modem7
        global COM7
        port = 'COM7'
        if action == 'Connect_device':
            try:
                modem7.smsTextMode = False
                modem7.connect()
                COM7 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if parametro[4] < 2:
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        date=[]
                        modem7.sendSms(str(parametro[1]), str(parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        date.append(str(time.asctime(time.localtime(time.time()))))
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':

                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F",Intentos = '+str(int(parametro[4]+1))+' WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM7 
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)
        elif action == 'Stop_device':
            try:
                modem7.close()
            except Exception as e:
                print(e)

    # Puerto 8
    def Com8(self, action):
        global modem8
        global COM8
        port = 'COM8'
        if action == 'Connect_device':
            try:
                modem8.smsTextMode = False
                modem8.connect()
                COM8 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if parametro[4] < 2:
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        date=[]
                        modem8.sendSms(str(parametro[1]), str(parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        date.append(str(time.asctime(time.localtime(time.time()))))
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':

                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F",Intentos = '+str(int(parametro[4]+1))+' WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM8 
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)
        elif action == 'Stop_device':
            try:
                modem8.close()
            except Exception as e:
                print(e)

    # Puerto 9
    def Com9(self, action):
        global modem9
        global COM9
        port = 'COM9'
        if action == 'Connect_device':
            try:
                modem9.smsTextMode = False
                modem9.connect()
                COM9 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if parametro[4] < 2:
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        date=[]
                        modem9.sendSms(str(parametro[1]), str(parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        date.append(str(time.asctime(time.localtime(time.time()))))
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':

                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F",Intentos = '+str(int(parametro[4]+1))+' WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM9 
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)
        elif action == 'Stop_device':
            try:
                modem9.close()
            except Exception as e:
                print(e)
    
    # Puerto 10
    def Com10(self, action):
        global modem10
        global COM10
        port = 'COM10'
        if action == 'Connect_device':
            try:
                modem10.smsTextMode = False
                modem10.connect()
                COM10 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if parametro[4] < 2:
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        date=[]
                        modem10.sendSms(str(parametro[1]), str(parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        date.append(str(time.asctime(time.localtime(time.time()))))
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':

                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F",Intentos = '+str(int(parametro[4]+1))+' WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM10 
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)
        elif action == 'Stop_device':
            try:
                modem10.close()
            except Exception as e:
                print(e)

    # Puerto 11
    def Com11(self, action):
        global modem11
        global COM11
        port = 'COM11'
        if action == 'Connect_device':
            try:
                modem11.smsTextMode = False
                modem11.connect()
                COM11 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if parametro[4] < 2:
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        date=[]
                        modem11.sendSms(str(parametro[1]), str(parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        date.append(str(time.asctime(time.localtime(time.time()))))
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':

                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F",Intentos = '+str(int(parametro[4]+1))+' WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM11 
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)
        elif action == 'Stop_device':
            try:
                modem11.close()
            except Exception as e:
                print(e)

    # Puerto 12
    def Com12(self, action):
        global modem12
        global COM12
        port = 'COM12'
        if action == 'Connect_device':
            try:
                modem12.smsTextMode = False
                modem12.connect()
                COM12 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':
            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if parametro[4] < 2:
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        date=[]
                        modem12.sendSms(str(parametro[1]), str(parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        date.append(str(time.asctime(time.localtime(time.time()))))
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio,Hora_ingreso_db) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(date[0])+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':

                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F",Intentos = '+str(int(parametro[4]+1))+' WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM12 
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)
        elif action == 'Stop_device':
            try:
                modem12.close()
            except Exception as e:
                print(e)

    # Abre el hilo principal de trabajo para el envio de mensajes de todos los puertos
    def start_send(self):
        if conected == True:
            if db_conn != '':
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setInformativeText(
                    'Iniciando envio de mensajes\nEste proceso puede demorar dependiendo de la cantidad de mensajes a enviar')
                msg.setWindowTitle("info")
                msg.exec_()
                hilostart = threading.Thread(
                    target=self.hilos, args=('Star_send',))
                hilostart.start()
                hilostart.join()
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setText("Info")
                msg.setInformativeText('Mensajes enviados con exito')
                msg.setWindowTitle("info")
                msg.exec_()
            else:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setInformativeText('Campaña no seleccionada')
                msg.setWindowTitle("Error")
                msg.exec_()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setInformativeText(
                'Por favor inicia los puertos para poder enviar mensajes')
            msg.setWindowTitle("Error")
            msg.exec_()

    # Consulta si aun hay mensajes existentes sin enviar para volver a encolarse.
    def SecondRound(self):
        con = sqlite3.connect(str(db_conn))
        cursor = con.cursor()
        again = False
        try:
            cursor.execute('SELECT * FROM SMS WHERE Enviado = "F"')
            rows = cursor.fetchall()
            if rows:
                again = True
        except Exception as e:
            print('rows second round: '+str(e))
        try:
            time.sleep(2.5)
            cursor.execute('SELECT * FROM SMS WHERE Enviado = "F"')
            rows = cursor.fetchall()
            if rows:
                again = True
        except Exception as e:
            print('rows second round: '+str(e))
        finally:
            if again == True:
                self.hilos('Star_send')

    # Funcion en la que los hilos correspondientes a cada puerto comienzan a trabajar
    def hilos(self, action):
        if action == 'Connect_device':
            com5 = threading.Thread(
                name='HCOM5', target=self.Com5, args=(action,))
            com6 = threading.Thread(
                name='HCOM6', target=self.Com6, args=(action,))
            com7 = threading.Thread(
                name='HCOM7', target=self.Com7, args=(action,))
            com8 = threading.Thread(
                name='HCOM8', target=self.Com8, args=(action,))
            com9 = threading.Thread(
                name='HCOM9', target=self.Com9, args=(action,))
            com10 = threading.Thread(
                name='HCOM10', target=self.Com10, args=(action,))
            com11 = threading.Thread(
                name='HCOM11', target=self.Com11, args=(action,))
            com12 = threading.Thread(
                name='HCOM12', target=self.Com12, args=(action,))
            com5.start()
            com6.start()
            com7.start()
            com8.start()
            com9.start()
            com10.start()
            com11.start()
            com12.start()
        elif action == 'Star_send':
            com5 = threading.Thread(
                name='HCOM5', target=self.Com5, args=(action,))
            com6 = threading.Thread(
                name='HCOM6', target=self.Com6, args=(action,))
            com7 = threading.Thread(
                name='HCOM7', target=self.Com7, args=(action,))
            com8 = threading.Thread(
                name='HCOM8', target=self.Com8, args=(action,))
            com9 = threading.Thread(
                name='HCOM9', target=self.Com9, args=(action,))
            com10 = threading.Thread(
                name='HCOM10', target=self.Com10, args=(action,))
            com11 = threading.Thread(
                name='HCOM11', target=self.Com11, args=(action,))
            com12 = threading.Thread(
                name='HCOM12', target=self.Com12, args=(action,))
            if COM5 == 'Is_Connected':
                com5.start()
            time.sleep(1)
            if COM6 == 'Is_Connected':
                com6.start()
            time.sleep(1)
            if COM7 == 'Is_Connected':
                com7.start()
            time.sleep(1)
            if COM8 == 'Is_Connected':
                com8.start()
            time.sleep(1)
            if COM9 == 'Is_Connected':
                com9.start()
            time.sleep(1)
            if COM10 == 'Is_Connected':
                com10.start()
            time.sleep(1)
            if COM11 == 'Is_Connected':
                com11.start()
            time.sleep(1)
            if COM12 == 'Is_Connected':
                com12.start()
            time.sleep(5)
            back = threading.Thread(target=self.SecondRound)
            back.start()
        elif action == 'Stop_device':
            com5 = threading.Thread(
                name='HCOM5', target=self.Com5, args=(action,))
            com6 = threading.Thread(
                name='HCOM6', target=self.Com6, args=(action,))
            com7 = threading.Thread(
                name='HCOM7', target=self.Com7, args=(action,))
            com8 = threading.Thread(
                name='HCOM8', target=self.Com8, args=(action,))
            com9 = threading.Thread(
                name='HCOM9', target=self.Com9, args=(action,))
            com10 = threading.Thread(
                name='HCOM10', target=self.Com10, args=(action,))
            com11 = threading.Thread(
                name='HCOM11', target=self.Com11, args=(action,))
            com12 = threading.Thread(
                name='HCOM12', target=self.Com12, args=(action,))
            com5.start()
            com6.start()
            com7.start()
            com8.start()
            com9.start()
            com10.start()
            com11.start()
            com12.start()

    # Menu lateral activa los index correspondientes a la seleccion.
    def MenuSend(self, it, col):
        if it.text(col) == 'Enviados':
            self.stackedWidget.setCurrentIndex(3)
        elif it.text(col) == 'Recibidos':
            self.stackedWidget.setCurrentIndex(4)
        elif it.text(col) == 'Base de datos':
            self.stackedWidget.setCurrentIndex(5)
        elif it.text(col) == 'Programados':
            self.stackedWidget.setCurrentIndex(2)

if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    window = MainWindow()
    window.show()
    app.exec_()