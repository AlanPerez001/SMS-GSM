from Firstqt_ui import *
import sys
import glob
import serial
from gsmmodem.modem import GsmModem
import serial.tools.list_ports
import time
import logging
from PyQt5.QtWidgets import QMessageBox, QFileDialog
import concurrent.futures
import threading
import sqlite3
import serial.tools.list_ports
import openpyxl

conected = False
db_conn = ''
COM5 = 'Is_Desconnected'
COM6 = 'Is_Desconnected'
COM7 = 'Is_Desconnected'
COM8 = 'Is_Desconnected'
COM9 = 'Is_Desconnected'
COM10 = 'Is_Desconnected'
COM11 = 'Is_Desconnected'
COM12 = 'Is_Desconnected'


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):

    # Inicia aplicacion y actionbuttons que activan las funciones
    def __init__(self, *args, **kwargs):
        global modem5
        global modem6
        global modem7
        global modem8
        global modem9
        global modem10
        global modem11
        global modem12
        modem5 = GsmModem(
            'COM5', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem6 = GsmModem(
            'COM6', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem7 = GsmModem(
            'COM7', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem8 = GsmModem(
            'COM8', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem9 = GsmModem(
            'COM9', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem10 = GsmModem(
            'COM10', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem11 = GsmModem(
            'COM11', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        modem12 = GsmModem(
            'COM12', 115200, smsReceivedCallbackFunc=self.ReceivedSms)
        QtWidgets.QMainWindow.__init__(self, *args, **kwargs)
        self.setupUi(self)
        self.stackedWidget.setCurrentIndex(1)
        self.progressBar.hide()
        self.progressBar_2.hide()
        self.treeWidget.itemClicked.connect(self.MenuSend)
        self.actionConectarModem.triggered.connect(self.actionConnect)
        self.actionEnviar_mensaje.triggered.connect(self.actionSend)
        self.actionProgramar_mensaje.triggered.connect(self.actionSchedul)
        self.actionExit.triggered.connect(self.actionExitapp)
        self.pushButton_5.clicked.connect(self.ConectModem)
        self.Actualizar.clicked.connect(self.ActualizarPort)
        self.ExitBTN.clicked.connect(self.actionExitapp)
        self.IniciarBTN.clicked.connect(self.startcon)
        self.Enviar.clicked.connect(self.start_send)
        self.comboBox.currentIndexChanged.connect(self.combo)
        self.PuertosBTN.clicked.connect(self.actionConnect)
        self.Connect_DB.clicked.connect(self.getfile)
        self.Actualiza_db.clicked.connect(self.updateDB)
        self.crear_campaa.clicked.connect(self.create_show)

   # Inicia la coneccion con todos los puertos disponibles
    def startcon(self):
        global conected
        hilos_conn = threading.Thread(
            target=self.hilos, args=('Connect_device',))
        hilos_conn.start()
        hilos_conn.join()
        conected = True
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setInformativeText('Conneccion establecida con los puertos :D')
        msg.setWindowTitle("info")
        msg.exec_()

    # Muestra el index 6
    def create_show(self):
        self.stackedWidget.setCurrentIndex(6)
        self.Importar_db.hide()
        self.Create_DB.clicked.connect(self.create_db)
        self.Importar_db.clicked.connect(self.import_db)

    # Lee los datos en el excel del usuario para insertarlos a la base de datos
    def import_db(self):
        excel_file, _filter = QFileDialog.getOpenFileName(self, 'Selecciona tu archivo XLSX a importar',
                                                          'C:\\Users\\Developer\\Desktop\\', "Excel File (*.xlsx)")
        print(excel_file)
        wb_obj = openpyxl.load_workbook(excel_file)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        con = sqlite3.connect(db_conn)
        cursor = con.cursor()
        for i in range(2, m_row + 1):
            numero = sheet_obj.cell(row=i, column=1)
            print(numero.value)
            mensaje = sheet_obj.cell(row=i, column=2)
            print(mensaje.value)
            cursor.execute('INSERT INTO SMS (Numero,Mensaje,Enviado) VALUES("' +
                           str(numero.value)+'","'+str(mensaje.value)+'","F")')
            con.commit()
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setInformativeText('Registros importados a la base de datos')
        msg.setWindowTitle("info")
        msg.exec_()
        self.updateDB()

    # crea una base de datos nueva con el nombre escrito por el usuario en la ubicacion seleccionada
    def create_db(self):
        global db_conn
        name_db = self.Name_create_db.toPlainText()
        files = QFileDialog.getExistingDirectoryUrl(
            None, 'C:\\Users\\Developer\\Desktop\\')
        path = str(files.path())
        path1 = path[1:len(path)].replace('/', '\\')
        database = str(path1)+'\\'+str(name_db)+'.db'
        sms = """ CREATE TABLE "SMS" (
                    "ID"	INTEGER UNIQUE,
                    "Numero"	INTEGER,
                    "Mensaje"	TEXT,
                    "Enviado"	TEXT,
                    PRIMARY KEY("ID" AUTOINCREMENT)
                    ); """

        enviados = """CREATE TABLE "Enviados" (
                            "ID"	INTEGER,
                            "Numero"	INTEGER,
                            "Mensaje"	TEXT,
                            "Puerto"	TEXT,
                            "Hora_envio"	TEXT,
                            PRIMARY KEY("ID" AUTOINCREMENT)
                            );"""

        recibidos = """CREATE TABLE "Recibidos" (
            "ID"	INTEGER UNIQUE,
            "Numero"	INTEGER,
            "Mensaje"	TEXT,
            "Hora_recibido"	TEXT,
            PRIMARY KEY("ID" AUTOINCREMENT)
            );"""
        conn = sqlite3.connect(database)
        try:
            c = conn.cursor()
            c.execute(sms)
        except Exception as e:
            print(e)
        try:
            c = conn.cursor()
            c.execute(enviados)
        except Exception as e:
            print(e)
        try:
            c = conn.cursor()
            c.execute(recibidos)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setInformativeText('Base de datos creada con exito :D')
            msg.setWindowTitle("info")
            msg.exec_()
            self.Importar_db.show()
            db_conn = database
        except Exception as e:
            print(e)

    # Obtiene la ruta de la base de datos seleccionada
    def getfile(self):
        global db_conn
        db_conn, _filter = QFileDialog.getOpenFileName(self, 'Selecciona tu base de datos',
                                                       'C:\\Users\\Developer\\Desktop\\', "data base file (*.db)")
        self.updateDB()

    # muestra el index 2 mensajes programados (Coming Soon)
    def actionSchedul(self):
        self.stackedWidget.setCurrentIndex(2)

    # Cierra conecciones con los puertos abiertos y cierra la app
    def actionExitapp(self):

        try:
            modem5.close()
            modem6.close()
            modem7.close()
            modem8.close()
            modem9.close()
            modem10.close()
            modem11.close()
            modem12.close()
            self.modem.close()
        except Exception as e:
            print(e)

        finally:
            app.exit()

    # Muestra index 1
    def actionConnect(self):
        self.stackedWidget.setCurrentIndex(1)

    # Muestra index 0
    def actionSend(self):
        self.stackedWidget.setCurrentIndex(0)
        self.pushButton.clicked.connect(self.indivudual_sms)

    # Devuelve y activa el index seleccionado en el Combobox
    def combo(self):
        seleccion = self.comboBox.currentText()
        if str(seleccion) == 'Componer':
            self.stackedWidget.setCurrentIndex(0)
            self.pushButton.clicked.connect(self.indivudual_sms)
        elif str(seleccion) == 'Programado':
            self.stackedWidget.setCurrentIndex(2)

    # Establece una coneccion individual del puerto seleccionado
    def ConectModem(self):
        try:
            portSelected = self.listWidget.currentItem().text()
            print(portSelected)
            try:
                self.modem = GsmModem(
                    str(portSelected), 115200, smsReceivedCallbackFunc=MainWindow.ReceivedSms)
                self.modem.smsTextMode = False
                self.modem.connect()
                print('Conectado')
            except Exception as F:
                print('El puerto '+str(portSelected) +
                      ' No esta disponible: ' + str(F))
        except Exception as e:
            print(e)

    # obtiene una lista de los puertos existentes
    def ActualizarPort(self):
        ports = serial.tools.list_ports.comports()
        advance = 0
        porcantaje = float(100/len(ports))
        self.listWidget.clear()
        num_hilo = 1
        for p in sorted(ports, key=None, reverse=False):
            advance += float(porcantaje)
            port = p.device
            try:
                self.progressBar.show()
                self.progressBar.setGeometry(QtCore.QRect(300, 330, 201, 21))
                self.progressBar.setProperty("value", advance)
                self.progressBar.setObjectName("progressBar")
                # Abre hilo llamando a la funcion Connectport para crear coneccion momentanea y corroborar el estado del puerto
                hilo = threading.Thread(name='hilo %s' % str(
                    num_hilo), target=self.connectport, args=(port,))
                hilo.start()
                num_hilo += 1
            finally:
                self.progressBar.hide()

    # Crea una coneccion momentanea para corroborar el estado de los puertos y  escribe los puertos en la lista desplegable
    def connectport(self, port):
        try:
            phone = serial.Serial(str(port),  115200, timeout=0.2)
            phone.write(b'ATZ\r')
            phone.write(b'AT+CMGF=1\r')
            if str(phone.read()) != str(b''):
                self.listWidget.addItem(port)
        except Exception as E:
            if str(E) == 'None':
                print('No disponibe :c')
        finally:
            try:
                phone.close()
            except Exception as f:
                print(f)

    # Envia mensaje de texto al numero escrito por el usuario por el primer puerto disponible
    def indivudual_sms(self):
        num = self.Num_edit.toPlainText()
        text = self.SMSText_edit.toPlainText()
        send = False
        if conected == True:
            if num != "" and text != "" and len(num) >= 10:
                while send == False:
                    if send == False:
                        try:
                            if COM5 == 'Is_Connected':
                                modem5.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM6 == 'Is_Connected':
                                modem6.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM7 == 'Is_Connected':
                                modem7.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM8 == 'Is_Connected':
                                modem8.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM9 == 'Is_Connected':
                                modem9.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM10 == 'Is_Connecte d':
                                modem10.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM11 == 'Is_Connected':
                                modem11.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                    if send == False:
                        try:
                            if COM12 == 'Is_Connected':
                                modem12.sendSms(str(num), str(
                                    text), waitForDeliveryReport=False, deliveryTimeout=1)
                                send = True
                                msg = QMessageBox()
                                msg.setIcon(QMessageBox.Information)
                                msg.setInformativeText(
                                    'Mensaje enviado con exito :D')
                                msg.setWindowTitle("info")
                                msg.exec_()
                                self.Num_edit.clear()
                                self.SMSText_edit.clear()
                        except Exception as F:
                            send = False
                            if str(F) == 'CMS 500':
                                print('El puerto seleccionado devolvio ' +
                                      str(F)+', [Error Unknown]')
                print('saliendo del ciclo')
            else:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setInformativeText(
                    'Por favor ingresa un numero y/o mesnsaje')
                msg.setWindowTitle("info")
                msg.exec_()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setInformativeText(
                'Por favor inicia los puertos para poder enviar mensajes')
            msg.setWindowTitle("Error")
            msg.exec_()

    # Cuando un mensaje se recibe redirige a la funcion para obtener el mensaje
    def ReceivedSms(self, sms):
        print(u'== SMS message received ==\nFrom: {0}\nTime: {1}\nMessage: {2}'.format(
            sms.number, sms.time, sms.text))
        con = sqlite3.connect(str(db_conn))
        cursor = con.cursor()
        cursor.execute('INSERT INTO Recibidos (Numero,Mensaje,Hora_recibido) VALUES("' +
                       str(sms.number)+'","'+str(sms.text)+'","'+str(sms.time)+'")')
        con.commit()
        self.updateDB()

    # Actualiza la base de datos Enviados y Mensajes
    def updateDB(self):
        if db_conn != '':
            con = sqlite3.connect(str(db_conn))
            cursor = con.cursor()
            cursor.execute('SELECT* FROM SMS')
            rows = cursor.fetchall()
            count = 0
            for i in rows:
                numero = QtWidgets.QTableWidgetItem()
                self.tableWidget_4.setItem(count, 0, numero)
                numero = self.tableWidget_4.item(count, 0)
                numero.setText(str(i[1]))
                mensaje = QtWidgets.QTableWidgetItem()
                self.tableWidget_4.setItem(count, 1, mensaje)
                mensaje = self.tableWidget_4.item(count, 1)
                mensaje.setText(str(i[2]))
                enviado = QtWidgets.QTableWidgetItem()
                self.tableWidget_4.setItem(count, 2, enviado)
                enviado = self.tableWidget_4.item(count, 2)
                enviado.setText(str(i[3]))
                count += 1

            cursor.execute('SELECT* FROM Enviados')
            rows = cursor.fetchall()
            count2 = 0
            for i in rows:
                numero = QtWidgets.QTableWidgetItem()
                self.tableWidget_3.setItem(count2, 0, numero)
                numero = self.tableWidget_3.item(count2, 0)
                numero.setText(str(i[1]))
                mensaje = QtWidgets.QTableWidgetItem()
                self.tableWidget_3.setItem(count2, 1, mensaje)
                mensaje = self.tableWidget_3.item(count2, 1)
                mensaje.setText(str(i[2]))
                com = QtWidgets.QTableWidgetItem()
                self.tableWidget_3.setItem(count2, 2, com)
                com = self.tableWidget_3.item(count2, 2)
                com.setText(str(i[3]))
                date = QtWidgets.QTableWidgetItem()
                self.tableWidget_3.setItem(count2, 3, date)
                date = self.tableWidget_3.item(count2, 3)
                date.setText(str(i[4]))
                count2 += 1

            cursor.execute('SELECT* FROM Recibidos')
            rows = cursor.fetchall()
            count3 = 0
            for i in rows:
                numero = QtWidgets.QTableWidgetItem()
                self.tableWidget_2.setItem(count3, 0, numero)
                numero = self.tableWidget_2.item(count3, 0)
                numero.setText(str(i[1]))
                mensaje = QtWidgets.QTableWidgetItem()
                self.tableWidget_2.setItem(count3, 1, mensaje)
                mensaje = self.tableWidget_2.item(count3, 1)
                mensaje.setText(str(i[2]))
                date = QtWidgets.QTableWidgetItem()
                self.tableWidget_2.setItem(count3, 2, date)
                date = self.tableWidget_2.item(count3, 2)
                date.setText(str(i[3]))
                count3 += 1
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Error")
            msg.setInformativeText(
                'Base de datos de campaña no seleccionados.')
            msg.setWindowTitle("Error")
            msg.exec_()

    # Puerto 5
    def Com5(self, action):
        global modem5
        global COM5
        port = 'COM5'
        if action == 'Connect_device':
            try:
                modem5.smsTextMode = False
                modem5.connect()
                COM5 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':
            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if str(parametro[3]) == 'F':
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        print('El hilo es: {0},   El parametro es: {1}'.format(
                            threading.current_thread().getName(), parametro[0]))
                        modem5.sendSms(str(parametro[1]), str(
                            parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':
                            print('No se logro enviar el mensaje ' +
                                  str(e)+' del puerto: '+port)
                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM5 = 'Is_Desconnected'
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)

    # Puerto 6
    def Com6(self, action):
        global modem6
        global COM6
        port = 'COM6'
        if action == 'Connect_device':
            try:
                print('Conectando puerto'+port)
                modem6.smsTextMode = False
                modem6.connect()
                COM6 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if str(parametro[3]) == 'F':
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        print('El hilo es: {0},   El parametro es: {1}'.format(
                            threading.current_thread().getName(), parametro[0]))
                        modem6.sendSms(str(parametro[1]), str(
                            parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':
                            print('No se logro enviar el mensaje ' +
                                  str(e)+' del puerto: '+port)
                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM6 = 'Is_Desconnected'
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)

    # Puerto 7
    def Com7(self, action):
        global modem7
        global COM7
        port = 'COM7'
        if action == 'Connect_device':
            try:
                print('Conectando puerto'+port)
                modem7.smsTextMode = False
                modem7.connect()
                COM7 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if str(parametro[3]) == 'F':
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        print('El hilo es: {0},   El parametro es: {1}'.format(
                            threading.current_thread().getName(), parametro[0]))
                        modem7.sendSms(str(parametro[1]), str(
                            parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':
                            print('No se logro enviar el mensaje ' +
                                  str(e)+' del puerto: '+port)
                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM7 = 'Is_Desconnected'
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)

    # Puerto 8
    def Com8(self, action):
        global modem8
        global COM8
        port = 'COM8'
        if action == 'Connect_device':
            try:
                print('Conectando puerto'+port)
                modem8.smsTextMode = False
                modem8.connect()
                COM8 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if str(parametro[3]) == 'F':
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        print('El hilo es: {0},   El parametro es: {1}'.format(
                            threading.current_thread().getName(), parametro[0]))
                        modem8.sendSms(str(parametro[1]), str(
                            parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':
                            print('No se logro enviar el mensaje ' +
                                  str(e)+' del puerto: '+port)
                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM8 = 'Is_Desconnected'
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)

    # Puerto 9
    def Com9(self, action):
        global modem9
        global COM9
        port = 'COM9'
        if action == 'Connect_device':
            try:
                print('Conectando puerto'+port)
                modem9.smsTextMode = False
                modem9.connect()
                COM9 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if str(parametro[3]) == 'F':
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        print('El hilo es: {0},   El parametro es: {1}'.format(
                            threading.current_thread().getName(), parametro[0]))
                        modem9.sendSms(str(parametro[1]), str(
                            parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':
                            print('No se logro enviar el mensaje ' +
                                  str(e)+' del puerto: '+port)
                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM9 = 'Is_Desconnected'
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)

    # Puerto 10
    def Com10(self, action):
        global modem10
        global COM10
        port = 'COM10'
        if action == 'Connect_device':
            try:
                print('Conectando puerto'+port)
                modem10.smsTextMode = False
                modem10.connect()
                COM10 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if str(parametro[3]) == 'F':
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        print('El hilo es: {0},   El parametro es: {1}'.format(
                            threading.current_thread().getName(), parametro[0]))
                        modem10.sendSms(str(parametro[1]), str(
                            parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':
                            print('No se logro enviar el mensaje ' +
                                  str(e)+' del puerto: '+port)
                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM10 = 'Is_Desconnected'
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)

    # Puerto 11
    def Com11(self, action):
        global modem11
        global COM11
        port = 'COM11'
        if action == 'Connect_device':
            try:
                print('Conectando puerto'+port)
                modem11.smsTextMode = False
                modem11.connect()
                COM11 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':

            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if str(parametro[3]) == 'F':
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        print('El hilo es: {0},   El parametro es: {1}'.format(
                            threading.current_thread().getName(), parametro[0]))
                        modem11.sendSms(str(parametro[1]), str(
                            parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':
                            print('No se logro enviar el mensaje ' +
                                  str(e)+' del puerto: '+port)
                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM11 = 'Is_Desconnected'
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)

    # Puerto 12
    def Com12(self, action):
        global modem12
        global COM12
        port = 'COM12'
        if action == 'Connect_device':
            try:
                print('Conectando puerto'+port)
                modem12.smsTextMode = False
                modem12.connect()
                COM12 = 'Is_Connected'
            except Exception as e:
                print(e)
        elif action == 'Star_send':
            try:
                con = sqlite3.connect(str(db_conn))
                cursor = con.cursor()
                cursor.execute('Select * FROM SMS WHERE Enviado ="F"')
                rows = cursor.fetchall()
                parametro = rows[0]
                if str(parametro[3]) == 'F':
                    cursor.execute(
                        'UPDATE SMS SET Enviado ="I" WHERE ID ='+str(parametro[0]))
                    con.commit()
                    try:
                        print('El hilo es: {0},   El parametro es: {1}'.format(
                            threading.current_thread().getName(), parametro[0]))
                        modem12.sendSms(str(parametro[1]), str(
                            parametro[2]), waitForDeliveryReport=False, deliveryTimeout=1)
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="T" WHERE ID ='+str(parametro[0]))
                        con.commit()
                        cursor.execute('INSERT INTO Enviados (Numero,Mensaje,Puerto,Hora_envio) VALUES("'+str(parametro[1])+'","'+str(
                            parametro[2])+'","'+str(port)+'","'+str(time.asctime(time.localtime(time.time())))+'")')
                        con.commit()
                    except Exception as e:
                        if str(e) == 'CMS 500':
                            print('No se logro enviar el mensaje ' +
                                  str(e)+' del puerto: '+port)
                            cursor.execute(
                                'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                            con.commit()
                            self.updateDB
                            COM12 = 'Is_Desconnected'
                        cursor.execute(
                            'UPDATE SMS SET Enviado ="F" WHERE ID ='+str(parametro[0]))
                        con.commit()
            except Exception as e:
                print('try de rows: '+str(e)+' del puerto: '+port)

    # Abre el hilo principal de trabajo para el envio de mensajes de todos los puertos
    def start_send(self):
        if conected == True:
            if db_conn != '':
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setInformativeText(
                    'Iniciando envio de mensajes\nEste proceso puede demorar dependiendo de la cantidad de mensajes a enviar')
                msg.setWindowTitle("info")
                msg.exec_()
                hilostart = threading.Thread(
                    target=self.hilos, args=('Star_send',))
                hilostart.start()
                hilostart.join()
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setText("Info")
                msg.setInformativeText('Mensajes enviados con exito')
                msg.setWindowTitle("info")
                msg.exec_()
            else:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setInformativeText('Campaña no seleccionada')
                msg.setWindowTitle("Error")
                msg.exec_()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setInformativeText(
                'Por favor inicia los puertos para poder enviar mensajes')
            msg.setWindowTitle("Error")
            msg.exec_()

    # Consulta si aun hay mensajes existentes sin enviar para volver a encolarse.
    def SecondRound(self):
        con = sqlite3.connect(str(db_conn))
        cursor = con.cursor()
        again = False
        try:
            cursor.execute('SELECT * FROM SMS WHERE Enviado = "F"')
            rows = cursor.fetchall()
            if rows:
                print('Si entro')
                for i in rows:
                    print(i)
                again = True
        except Exception as e:
            print('rows second round: '+str(e))
        try:
            cursor.execute('SELECT * FROM SMS WHERE Enviado = "I"')
            rows = cursor.fetchall()
            if rows:
                print('Si hilos intentando enviar mensajes')
                for i in rows:
                    print(i)
                again = True
                time.sleep(5)
                cursor.execute('SELECT * FROM SMS WHERE Enviado = "F"')
                rows = cursor.fetchall()
                if rows:
                    print('Si entro')
        except Exception as e:
            print('rows second round: '+str(e))
        finally:
            if again == True:
                self.hilos('Star_send')

    # Funcion en la que los hilos correspondientes a cada puerto comienzan a trabajar
    def hilos(self, action):
        if action == 'Connect_device':
            com5 = threading.Thread(
                name='HCOM5', target=self.Com5, args=(action,))
            com6 = threading.Thread(
                name='HCOM6', target=self.Com6, args=(action,))
            com7 = threading.Thread(
                name='HCOM7', target=self.Com7, args=(action,))
            com8 = threading.Thread(
                name='HCOM8', target=self.Com8, args=(action,))
            com9 = threading.Thread(
                name='HCOM9', target=self.Com9, args=(action,))
            com10 = threading.Thread(
                name='HCOM10', target=self.Com10, args=(action,))
            com11 = threading.Thread(
                name='HCOM11', target=self.Com11, args=(action,))
            com12 = threading.Thread(
                name='HCOM12', target=self.Com12, args=(action,))
            com5.start()
            com6.start()
            com7.start()
            com8.start()
            com9.start()
            com10.start()
            com11.start()
            com12.start()
        elif action == 'Star_send':
            com5 = threading.Thread(
                name='HCOM5', target=self.Com5, args=(action,))
            com6 = threading.Thread(
                name='HCOM6', target=self.Com6, args=(action,))
            com7 = threading.Thread(
                name='HCOM7', target=self.Com7, args=(action,))
            com8 = threading.Thread(
                name='HCOM8', target=self.Com8, args=(action,))
            com9 = threading.Thread(
                name='HCOM9', target=self.Com9, args=(action,))
            com10 = threading.Thread(
                name='HCOM10', target=self.Com10, args=(action,))
            com11 = threading.Thread(
                name='HCOM11', target=self.Com11, args=(action,))
            com12 = threading.Thread(
                name='HCOM12', target=self.Com12, args=(action,))
            if COM5 == 'Is_Connected':
                com5.start()
            time.sleep(1)
            if COM6 == 'Is_Connected':
                com6.start()
            time.sleep(1)
            if COM7 == 'Is_Connected':
                com7.start()
            time.sleep(1)
            if COM8 == 'Is_Connected':
                com8.start()
            time.sleep(1)
            if COM9 == 'Is_Connected':
                com9.start()
            time.sleep(1)
            if COM10 == 'Is_Connected':
                com10.start()
            time.sleep(1)
            if COM11 == 'Is_Connected':
                com11.start()
            time.sleep(1)
            if COM12 == 'Is_Connected':
                com12.start()
            time.sleep(5)
            back = threading.Thread(target=self.SecondRound)
            back.start()

    # Menu lateral activa los index correspondientes a la seleccion.
    def MenuSend(self, it, col):
        if it.text(col) == 'Enviados':
            print('Entrando a Enviados')
            self.stackedWidget.setCurrentIndex(3)
        elif it.text(col) == 'Recibidos':
            print('Entrando a Recibidos')
            self.stackedWidget.setCurrentIndex(4)
        elif it.text(col) == 'Base de datos':
            print('Entrando a Base')
            self.stackedWidget.setCurrentIndex(5)
        elif it.text(col) == 'Programados':
            print('Entrando a Programados')
            self.stackedWidget.setCurrentIndex(2)

if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    window = MainWindow()
    window.show()
    app.exec_()
